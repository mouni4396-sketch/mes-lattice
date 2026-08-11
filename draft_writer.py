"""
draft_writer.py  —  Agent 6 (Intake) stage 3: draft Excel writer.

DETERMINISTIC. No AI. Takes the mapper's validated JSON rows and writes them
into a COPY of MES-Overlay-TEMPLATE.xlsx, placing each row in the correct sheet
and column. Adds nothing the model didn't produce except:
  - a derived IRI local name (camelCase / PascalCase of the name)
  - the vendor prefix in the Layer column
  - an amber highlight on rows with confidence < REVIEW_THRESHOLD, so the human
    reviewer's attention goes where the model was least sure.

Output is the reviewable artifact: MES-<Vendor>-Overlay-DRAFT.xlsx.
It is a DRAFT for human review, never loaded to the graph directly.

Usage:
    from draft_writer import write_draft
    path = write_draft(mapper_result["rows"], vendor="opc",
                       template_path="MES-Overlay-TEMPLATE.xlsx",
                       out_path="MES-OPC-Overlay-DRAFT.xlsx")
"""

import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

REVIEW_THRESHOLD = 0.6
LOW_FILL = PatternFill("solid", fgColor="FFE9B8")   # amber = check this row

SHEET_NODES = "1. Nodes"
SHEET_ATTRS = "2. Data Properties"
SHEET_OBJS = "3. Object Properties"


def _iri_local(name: str, pascal: bool) -> str:
    """Make an IRI local name from a display name. Node classes -> PascalCase,
    properties/attributes -> camelCase."""
    if not name:
        return ""
    parts = re.split(r"[^A-Za-z0-9]+", name.strip())
    parts = [p for p in parts if p]
    if not parts:
        return ""
    if pascal:
        return "".join(p[:1].upper() + p[1:] for p in parts)
    first = parts[0]
    rest = parts[1:]
    return first[:1].lower() + first[1:] + "".join(p[:1].upper() + p[1:] for p in rest)


def _find_headers(ws):
    return {ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=c).value}


def _next_row(ws):
    """First empty data row (row 1 = headers, row 2 may be the example)."""
    r = 2
    # skip the greyed example row if present (col1 has italic example text)
    while ws.cell(row=r, column=1).value not in (None, ""):
        r += 1
    return r


def _put(ws, row, headers, colname, value):
    c = headers.get(colname)
    if c and value not in (None, ""):
        ws.cell(row=row, column=c, value=value)


def _flag_low(ws, row, headers, confidence):
    if confidence is not None and confidence < REVIEW_THRESHOLD:
        cc = headers.get("Confidence")
        if cc:
            ws.cell(row=row, column=cc).fill = LOW_FILL
        nc = headers.get("Name") or headers.get("Attribute") or headers.get("Property")
        if nc:
            ws.cell(row=row, column=nc).fill = LOW_FILL


def _delete_example_row(ws):
    """Remove the greyed template example row (row 2) if it looks like one."""
    # heuristic: the template example row 1st cell is a layer token like 'cm'
    # and the row is styled italic/grey. We simply clear row 2 if it's the
    # documented example (it always sits at row 2 in the shipped template).
    if ws.max_row >= 2:
        # detect: example rows were written with italic font by the template
        cell = ws.cell(row=2, column=1)
        font = cell.font
        if font and font.italic:
            ws.delete_rows(2, 1)


def write_draft(rows, vendor, template_path, out_path):
    """
    rows: list of mapper JSON objects, each:
      {"concept": {...}, "attributes": [...], "relationships": [...]}
    vendor: layer token, e.g. "opc"
    Returns out_path.
    """
    template_path = Path(template_path)
    out_path = Path(out_path)
    shutil.copyfile(template_path, out_path)

    wb = load_workbook(out_path)
    ws_nodes = wb[SHEET_NODES]
    ws_attrs = wb[SHEET_ATTRS]
    ws_objs = wb[SHEET_OBJS]

    for ws in (ws_nodes, ws_attrs, ws_objs):
        _delete_example_row(ws)

    h_nodes = _find_headers(ws_nodes)
    h_attrs = _find_headers(ws_attrs)
    h_objs = _find_headers(ws_objs)

    for obj in rows:
        c = obj.get("concept", {}) or {}
        cname = c.get("name", "")
        ctype = c.get("type", "")
        csource = c.get("source", "")

        # ---- 1. Nodes ----
        r = _next_row(ws_nodes)
        _put(ws_nodes, r, h_nodes, "Layer", vendor)
        _put(ws_nodes, r, h_nodes, "Name", cname)
        _put(ws_nodes, r, h_nodes, "IRI local name",
             _iri_local(cname, pascal=True))
        _put(ws_nodes, r, h_nodes, "Type", ctype)
        _put(ws_nodes, r, h_nodes, "Maps to neutral", c.get("maps_to_neutral", ""))
        _put(ws_nodes, r, h_nodes, "Match type", c.get("match_type", ""))
        _put(ws_nodes, r, h_nodes, "Verdict", c.get("verdict", ""))
        _put(ws_nodes, r, h_nodes, "Confidence", c.get("confidence", ""))
        _put(ws_nodes, r, h_nodes, "Source doc", csource)
        _put(ws_nodes, r, h_nodes, "Notes", c.get("notes", ""))
        _flag_low(ws_nodes, r, h_nodes, c.get("confidence"))

        # ---- 2. Data Properties ----
        for a in obj.get("attributes", []) or []:
            r = _next_row(ws_attrs)
            _put(ws_attrs, r, h_attrs, "Layer", vendor)
            _put(ws_attrs, r, h_attrs, "Data object", cname)
            _put(ws_attrs, r, h_attrs, "Attribute", a.get("name", ""))
            _put(ws_attrs, r, h_attrs, "IRI local name",
                 _iri_local(a.get("name", ""), pascal=False))
            _put(ws_attrs, r, h_attrs, "Datatype", a.get("datatype", ""))
            _put(ws_attrs, r, h_attrs, "Value kind", a.get("value_kind", ""))
            _put(ws_attrs, r, h_attrs, "Maps to neutral attribute",
                 a.get("maps_to_neutral", ""))
            _put(ws_attrs, r, h_attrs, "Match type", a.get("match_type", ""))
            _put(ws_attrs, r, h_attrs, "Verdict", a.get("verdict", ""))
            _put(ws_attrs, r, h_attrs, "Confidence", a.get("confidence", ""))
            _put(ws_attrs, r, h_attrs, "Source doc", a.get("source", csource))
            _put(ws_attrs, r, h_attrs, "Notes", a.get("notes", ""))
            _flag_low(ws_attrs, r, h_attrs, a.get("confidence"))

        # ---- 3. Object Properties ----
        for rel in obj.get("relationships", []) or []:
            r = _next_row(ws_objs)
            _put(ws_objs, r, h_objs, "Layer", vendor)
            _put(ws_objs, r, h_objs, "Property", rel.get("name", ""))
            _put(ws_objs, r, h_objs, "IRI local name",
                 _iri_local(rel.get("name", ""), pascal=False))
            _put(ws_objs, r, h_objs, "Domain", cname)
            _put(ws_objs, r, h_objs, "Range", rel.get("target", ""))
            _put(ws_objs, r, h_objs, "Maps to neutral", rel.get("maps_to_neutral", ""))
            _put(ws_objs, r, h_objs, "Match type", rel.get("match_type", ""))
            _put(ws_objs, r, h_objs, "Verdict", rel.get("verdict", ""))
            _put(ws_objs, r, h_objs, "Confidence", rel.get("confidence", ""))
            _put(ws_objs, r, h_objs, "Source doc", rel.get("source", csource))
            _put(ws_objs, r, h_objs, "Notes", rel.get("notes", ""))
            _flag_low(ws_objs, r, h_objs, rel.get("confidence"))

    wb.save(out_path)
    return str(out_path)


if __name__ == "__main__":
    # smoke test with a synthetic mapper row
    demo = [{
        "concept": {"name": "Workflow", "type": "Data Object",
                    "maps_to_neutral": "Route", "match_type": "closeMatch",
                    "verdict": "covered", "confidence": 0.9,
                    "source": "docs/wf", "notes": "routing container"},
        "attributes": [
            {"name": "Workflow Name", "datatype": "string", "value_kind": "plain",
             "maps_to_neutral": "name", "match_type": "closeMatch",
             "verdict": "covered", "confidence": 0.95, "source": "docs/wf"},
            {"name": "Some Vague Field", "datatype": "string", "value_kind": "plain",
             "maps_to_neutral": "", "match_type": "", "verdict": "needs-verification",
             "confidence": 0.4, "source": "docs/wf"},
        ],
        "relationships": [
            {"name": "hasStage", "target": "Stage", "maps_to_neutral": "includesStep",
             "match_type": "broadMatch", "verdict": "partial", "confidence": 0.7,
             "source": "docs/wf"},
        ],
    }]
    out = write_draft(demo, vendor="opc",
                      template_path="MES-Overlay-TEMPLATE.xlsx",
                      out_path="MES-OPC-Overlay-DRAFT.xlsx")
    print("wrote", out)
