"""
generate_ttl.py — vendor-agnostic Excel -> OWL/TTL generator.

Turns a filled capture workbook into a {vendor}: overlay TTL that imports the frozen
ref: layer and emits ONLY vendor deltas. Excel is master; TTL is generated.

DESIGN RULES (why this file looks the way it does)
  A. Columns are read BY HEADER NAME, never by position. Vendor masters drift
     (Opcenter dropped "Confidence"); positional reads silently shift every later
     column. A missing REQUIRED header fails loudly; optional headers are genuinely
     optional (absent = omit the triple, never shift).
  B. Every literal is quoted and escaped. Confidence is emitted only when present
     AND numeric.
  C. Attribute/property IRIs are auto-qualified by owning class ONLY when the
     hand-authored local name would collide. CM masters already hand-qualify
     (areaName, batchName) and those are preserved verbatim.
  D. The verb-axis meta-class is namespaced so it cannot collide with a vendor noun
     (Camstar has "Operation" as a data object). Build fails if any meta IRI collides
     with an IRI minted from the master.
  E. A "Promotions" sheet (vendor independently confirms a Derived ref: concept) is
     emitted; it used to be silently dropped.
  F. Lookup allowed values and object-property Notes are carried through.
  G. A validation pass runs after generation: re-parses the TTL, checks for duplicate
     IRI definitions, mapping targets outside ref:, ref: IRIs used as subjects, and
     datatype properties whose range is a class. Prints a per-section count table.

Usage:
    python generate_ttl.py MASTER.xlsx --prefix opc --out mes-opc-overlay.ttl
"""

import argparse
import re
import sys
import openpyxl

BASE = "https://ontology.yourorg.com/mes-"
REF_NS = f"{BASE}ref#"

DTYPE = {"string": "xsd:string", "decimal": "xsd:decimal", "integer": "xsd:integer",
         "boolean": "xsd:boolean", "datetime": "xsd:dateTime"}
MATCH_OK = {"closematch", "broadmatch", "narrowmatch", "relatedmatch"}
# canonical SKOS spelling from a lowercased key
MATCH_CANON = {"closematch": "closeMatch", "broadmatch": "broadMatch",
               "narrowmatch": "narrowMatch", "relatedmatch": "relatedMatch"}
VERDICT_OK = {"covered", "partial", "vendor-extension", "absent", "needs-verification"}

# ---------------------------------------------------------------------------
# MES-Overlay-TEMPLATE.xlsx contract (new template). The legacy CM master
# (MES-CM-Overlay-MASTER.xlsx) predates this contract: no "0. Guide" version
# marker, sheets named "1. Data Objects"/"2. Attributes" instead of "1. Nodes"/
# "2. Data Properties", no context sheets, no Layer column on sheets 2-3.
# Every check below treats that shape as legacy and degrades gracefully.
# ---------------------------------------------------------------------------
EXPECTED_TEMPLATE_VERSION = "1.0"


# ---------------------------------------------------------------------------
# Header-name column access (fix A)
# ---------------------------------------------------------------------------
class Sheet:
    """Wraps a worksheet with case-insensitive, whitespace-tolerant header lookup."""

    def __init__(self, wb, name, required=(), optional=()):
        if name not in wb.sheetnames:
            raise BuildError(f"Workbook is missing required sheet '{name}'. "
                             f"Sheets present: {wb.sheetnames}")
        self.name = name
        self.ws = wb[name]
        raw = [c.value for c in next(self.ws.iter_rows(min_row=1, max_row=1))]
        self.headers = {}
        for i, h in enumerate(raw):
            if h is None:
                continue
            self.headers[str(h).strip().lower()] = i
        missing = [h for h in required if h.lower() not in self.headers]
        if missing:
            raise BuildError(
                f"Sheet '{name}' is missing required column(s): {missing}. "
                f"Found: {sorted(self.headers)}")
        self.optional = {o.lower() for o in optional}

    def has(self, header):
        return header.strip().lower() in self.headers

    def rows(self):
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            yield RowView(row, self.headers)

    def rows_with_index(self):
        """Like rows(), but also yields the 1-based Excel row number, for validation
        messages that must name the exact row (min_row=2 => first data row is 2)."""
        for i, row in enumerate(self.ws.iter_rows(min_row=2, values_only=True), start=2):
            yield i, RowView(row, self.headers)


def sheet_header_names(wb, name):
    """Lowercased header set of a sheet, read without enforcing any required columns -
    used to decide (before opening a Sheet, which does enforce) whether an optional-in-
    legacy column like 'Layer' is actually present on this workbook's copy of a sheet
    whose NAME is unchanged between the old and new template (e.g. '3. Object Properties')."""
    if name not in wb.sheetnames:
        return set()
    row1 = next(wb[name].iter_rows(min_row=1, max_row=1))
    return {str(c.value).strip().lower() for c in row1 if c.value is not None}


def open_sheet(wb, names, required=(), legacy_required=None, optional=()):
    """
    Open a sheet trying each candidate name in order - the new template's name first,
    then the legacy CM-master name - so a workbook in either shape works unmodified.
    `names` may be a single sheet name or a tuple of aliases.

    `legacy_required`, if given, is the required-header set used when the LEGACY name
    resolved (older masters have fewer columns, e.g. no "Layer" on sheets 2-3); when the
    new name resolved, `required` (the full new-template contract) applies.
    """
    if isinstance(names, str):
        names = (names,)
    present = [n for n in names if n in wb.sheetnames]
    if not present:
        raise BuildError(
            f"Workbook is missing sheet {names[0]!r}"
            + (f" (also tried {list(names[1:])})" if len(names) > 1 else "") +
            f". Sheets present: {wb.sheetnames}")
    resolved = present[0]
    is_legacy_name = legacy_required is not None and resolved != names[0]
    req = legacy_required if is_legacy_name else required
    return Sheet(wb, resolved, required=req, optional=optional), resolved


# ---------------------------------------------------------------------------
# Row validation (section 3). One shared checker for sheets 1-3: each has a
# Layer, a name-ish column (Name/Attribute/Property), an optional IRI local
# name fallback, and the same Match type / VERDICT / Confidence / secondary-
# mapping shape - only the column names differ per sheet.
# ---------------------------------------------------------------------------
def check_row_problems(sheet_label, row_num, r, sh, name_col, mapping_col,
                        match_col="Match type", verdict_col="VERDICT",
                        sec_mapping_col="Secondary maps to neutral",
                        sec_match_col="Secondary match type", warnings=None):
    problems = []
    if sh.has("Layer") and not r.str("Layer"):
        problems.append(f"{sheet_label} row {row_num}: Layer is empty.")

    name_val = r.str(name_col)
    iri_val = r.str("IRI local name") if sh.has("IRI local name") else ""
    if not name_val and not iri_val:
        problems.append(
            f"{sheet_label} row {row_num}: both '{name_col}' and 'IRI local name' are empty.")

    match_val = r.str(match_col) if sh.has(match_col) else ""
    if match_val and match_val.lower().replace(" ", "") not in MATCH_OK:
        problems.append(
            f"{sheet_label} row {row_num}, {match_col}: {match_val!r} is not one of "
            f"closeMatch, broadMatch, narrowMatch, relatedMatch.")

    verdict_val = r.str(verdict_col) if sh.has(verdict_col) else ""
    if verdict_val and verdict_val.strip().lower() not in VERDICT_OK:
        problems.append(
            f"{sheet_label} row {row_num}, {verdict_col}: {verdict_val!r} is not one of "
            f"{sorted(VERDICT_OK)}.")

    conf_val = r.str("Confidence") if sh.has("Confidence") else ""
    if conf_val:
        d = as_decimal(conf_val)
        if d is None or not (0.0 <= float(d) <= 1.0):
            problems.append(
                f"{sheet_label} row {row_num}, Confidence: {conf_val!r} does not parse "
                f"as a decimal in [0,1].")

    mapping_val = r.str(mapping_col) if sh.has(mapping_col) else ""
    if mapping_val and not match_val and warnings is not None:
        warnings.append(
            f"WARNING: {sheet_label} row {row_num}: '{mapping_col}' is set but "
            f"{match_col} is empty.")

    sec_tgt = r.str(sec_mapping_col) if sh.has(sec_mapping_col) else ""
    sec_match = r.str(sec_match_col) if sh.has(sec_match_col) else ""
    if sec_tgt and not sec_match:
        problems.append(
            f"{sheet_label} row {row_num}: '{sec_mapping_col}' is set but "
            f"'{sec_match_col}' is empty.")
    elif sec_match and sec_match.lower().replace(" ", "") not in MATCH_OK:
        problems.append(
            f"{sheet_label} row {row_num}, {sec_match_col}: {sec_match!r} is not one of "
            f"closeMatch, broadMatch, narrowMatch, relatedMatch.")
    return problems


class RowView:
    def __init__(self, row, headers):
        self._row, self._h = row, headers

    def get(self, header, default=""):
        i = self._h.get(header.strip().lower())
        if i is None or i >= len(self._row):
            return default
        v = self._row[i]
        return default if v is None else v

    def str(self, header):
        return str(self.get(header, "")).strip()

    def is_blank(self):
        """True for a spacer/fully-empty Excel row - never a validation target."""
        return all(v is None or str(v).strip() == "" for v in self._row)


class BuildError(Exception):
    pass


# ---------------------------------------------------------------------------
# Literal handling (fix B)
# ---------------------------------------------------------------------------
def lit(value):
    """Quote+escape a Turtle string literal."""
    s = str(value).replace("\\", "\\\\").replace('"', '\\"')
    s = s.replace("\r", " ").replace("\n", " ").strip()
    return f'"{s}"'


def as_decimal(value):
    """Return a Turtle decimal if value parses as a number, else None."""
    try:
        f = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    return repr(f) if f != int(f) else str(int(f))


def norm(value):
    """
    Normalise a cell value for MATCHING (not for output): strip whitespace, strip any
    wrapping quote characters, collapse internal whitespace, lowercase.
    Masters sometimes carry values like "'opc'" (literal quotes inside the cell), which
    silently failed an exact comparison and skipped every row.
    """
    s = str(value).strip()
    s = s.strip("'\"\u2018\u2019\u201c\u201d")
    s = re.sub(r"\s+", " ", s)
    return s.strip().lower()


def local(name):
    """Turn a display name or IRI into a valid IRI local part."""
    s = str(name).strip()
    if "#" in s:
        s = s.rsplit("#", 1)[1]
    if ":" in s and " " not in s:
        s = s.split(":", 1)[1]
    if " " in s:
        parts = s.split()
        s = parts[0][:1].lower() + parts[0][1:] + "".join(p[:1].upper() + p[1:] for p in parts[1:])
    return re.sub(r"[^A-Za-z0-9_\-]", "", s)


def cls_local(name):
    """Class-style local name (leading capital)."""
    s = local(name)
    return s[:1].upper() + s[1:] if s else s


def ref_iri(value):
    s = str(value).strip()
    if s.startswith("ref:"):
        return s
    return "ref:" + local(s)


def yes_no(value):
    """"Is key"/"Is mandatory" style cells: 'Yes' -> true, anything else (incl. blank) -> false."""
    return "true" if str(value).strip().lower() == "yes" else "false"


# ---------------------------------------------------------------------------
# Generator
# ---------------------------------------------------------------------------
class Generator:
    def __init__(self, path, prefix):
        self.wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        self.p = prefix
        self.ns = f"{BASE}{prefix}#"
        self.out = []
        self.defined = {}          # local name -> section (duplicate detection, fix C/G)
        self.name_to_iri = {}      # display name (normalised) -> minted class IRI
        self.warnings = []
        self.counts = {}
        self.is_legacy = True      # set by check_template_version(); no "0. Guide" => legacy
        self.field_iris = {}       # (table name, field display name) -> field IRI, for section 6
        # fix D: verb-axis meta-class is namespaced so a vendor noun ("Operation")
        # can never collide with it.
        self.OP_CLASS = "PortalOperation"
        self.meta_iris = {
            "DataObject", "ExternalReference", self.OP_CLASS,
            "operatesOn", "verdict", "sourceDoc", "operationKind", "confidence",
            "confirmsRefConcept", "promotionNote", "promotionEvidence", "UnresolvedReference",
            # context sheets (section 6)
            "ContextResolver", "ContextField", "PrecedenceTier", "resolvesTarget",
            "tableKind", "ofTable", "isKey", "isMandatory", "fieldComment",
            "keyField", "tierOrder",
        }

    def w(self, s=""):
        self.out.append(s)

    def define(self, name, section):
        """
        Register an IRI local name.

        Hard failure (fix C/G) for collisions that corrupt meaning: two properties, or
        a property colliding with a class. Multiple rdfs:domain on one property is
        CONJUNCTIVE in RDFS, so a duplicated property IRI silently entails that anything
        carrying it belongs to every domain class at once.

        A class/enum-set name overlap (e.g. a TransferRequirementType data object and a
        TransferRequirementType enum) is recorded as a warning instead: it is a naming
        smell worth surfacing, but the two are different node types and do not corrupt
        entailment.
        """
        prior = self.defined.get(name)
        if prior:
            typed = {"1. Data Objects": "class", "4. Enums & Lookups": "class",
                     "5. Operations": "class"}
            kind_a, kind_b = typed.get(prior, "property"), typed.get(section, "property")
            if kind_a == "class" and kind_b == "class":
                self.warnings.append(
                    f"name reused across sections: {self.p}:{name} "
                    f"({prior} and {section})")
                return
            raise BuildError(
                f"IRI {self.p}:{name} is defined twice "
                f"(first in {prior}, again in {section}). "
                f"Qualify the 'IRI local name' in the master to disambiguate.")
        self.defined[name] = section

    # ---- template version check (section 1) ----
    def check_template_version(self):
        """
        Scan '0. Guide' for a cell reading 'Template Version' and read the cell to its
        right. Absent sheet, absent label, or an empty value cell => legacy CM master;
        skip silently (self.is_legacy stays True). A present-but-mismatched version is
        a warning, never fatal.
        """
        if "0. Guide" not in self.wb.sheetnames:
            return
        ws = self.wb["0. Guide"]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip().lower() == "template version":
                    right = ws.cell(row=cell.row, column=cell.column + 1).value
                    version = str(right).strip() if right is not None else ""
                    if not version:
                        return
                    self.is_legacy = False
                    if version != EXPECTED_TEMPLATE_VERSION:
                        self.warnings.append(
                            f"WARNING: '0. Guide' declares Template Version {version!r}, "
                            f"this generator expects {EXPECTED_TEMPLATE_VERSION!r}.")
                    return

    # ---- row validation (section 3) ----
    def validate_rows(self):
        """
        Validate every data row on sheets 1-3, across ALL vendors present in the
        workbook (not just this build's --prefix) - a bad row for another vendor must
        not be silently ignored just because this build isn't emitting that vendor's
        overlay. Collects every problem, then raises ONE BuildError naming them all;
        the caller (generate()) runs this before writing a single line of TTL, so a
        partial file is never produced.
        """
        problems = []

        sh1, label1 = open_sheet(
            self.wb, ("1. Nodes", "1. Data Objects"),
            required=["Layer", "Name", "IRI local name", "Type"],
            optional=["Maps to neutral", "Match type", "VERDICT",
                      "Secondary maps to neutral", "Secondary match type", "Confidence"])
        for i, r in sh1.rows_with_index():
            if r.is_blank():
                continue
            problems += check_row_problems(label1, i, r, sh1, "Name", "Maps to neutral",
                                            warnings=self.warnings)

        sh2, label2 = open_sheet(
            self.wb, ("2. Data Properties", "2. Attributes"),
            required=["Data object", "Attribute", "IRI local name", "Datatype"],
            optional=["Layer", "Maps to neutral attribute", "Match type", "VERDICT",
                      "Secondary maps to neutral", "Secondary match type", "Confidence"])
        for i, r in sh2.rows_with_index():
            if r.is_blank():
                continue
            problems += check_row_problems(label2, i, r, sh2, "Attribute",
                                            "Maps to neutral attribute", warnings=self.warnings)

        has_layer_col = "layer" in sheet_header_names(self.wb, "3. Object Properties")
        sh3 = Sheet(self.wb, "3. Object Properties",
                    required=["Domain", "Property", "IRI local name", "Range"],
                    optional=(["Layer"] if has_layer_col else [])
                             + ["Maps to neutral", "Match type", "VERDICT",
                                "Secondary maps to neutral", "Secondary match type", "Confidence"])
        for i, r in sh3.rows_with_index():
            if r.is_blank():
                continue
            problems += check_row_problems("3. Object Properties", i, r, sh3, "Property",
                                            "Maps to neutral", warnings=self.warnings)

        if problems:
            raise BuildError("Row validation failed - " + str(len(problems)) +
                              " problem(s):\n  " + "\n  ".join(problems))

    # ---- shared verdict + mapping-edge emission (sections 4 & 5) ----
    def verdict_for(self, sh, r, mapping_col="Maps to neutral", match_col="Match type",
                     verdict_col="VERDICT"):
        """
        The row's own Verdict wins when the column exists and the cell is non-empty.
        Otherwise derive it (fix: backward-compat default for masters/rows that never
        carried a Verdict at all): mapped + closeMatch -> covered; mapped + broad/narrow/
        related -> partial; not mapped -> vendor-extension. Returns None (omit the triple)
        when a mapping target is given but the match type is missing/unrecognised - that
        combination is already flagged as a (non-fatal) row-validation warning.
        Callers only reach this after filtering out Layer == 'ref' rows, so the "ref-layer
        rows never get a verdict" rule is satisfied by construction, not checked here.

        If the sheet carries NEITHER a Verdict NOR a mapping column at all (e.g. the legacy
        CM master's "2. Attributes", which predates verdict/mapping on that sheet entirely),
        this returns None unconditionally rather than defaulting every row to
        "vendor-extension" - defaulting is for a sheet that HAS the concept and a blank
        cell, not for a sheet that never had the concept.
        """
        has_verdict_col = sh.has(verdict_col)
        has_mapping_col = sh.has(mapping_col)
        if not has_verdict_col and not has_mapping_col:
            return None
        if has_verdict_col:
            explicit = r.str(verdict_col)
            if explicit:
                return explicit
        has_mapping = has_mapping_col and bool(r.str(mapping_col))
        if not has_mapping:
            return "vendor-extension"
        m = r.str(match_col).lower().replace(" ", "") if sh.has(match_col) else ""
        if m == "closematch":
            return "covered"
        if m in ("broadmatch", "narrowmatch", "relatedmatch"):
            return "partial"
        return None

    def emit_mapping_edges(self, sh, r, iri, mapping_col="Maps to neutral",
                            match_col="Match type",
                            sec_mapping_col="Secondary maps to neutral",
                            sec_match_col="Secondary match type"):
        """Emit the primary SKOS mapping edge (existing behaviour) plus, when present, a
        SECOND edge to 'Secondary maps to neutral' (section 5). A secondary target with a
        missing/invalid match type is a hard row-validation error caught before generation
        ever reaches this point, so it is safe to just skip emitting it here."""
        p = self.p
        if sh.has(mapping_col) and sh.has(match_col):
            m = r.str(match_col).lower().replace(" ", "")
            tgt = r.str(mapping_col)
            if tgt and m in MATCH_OK:
                self.w(f"{p}:{iri} skos:{MATCH_CANON[m]} {ref_iri(tgt)} .")
        if sh.has(sec_mapping_col):
            sec_tgt = r.str(sec_mapping_col)
            if sec_tgt and sh.has(sec_match_col):
                sm = r.str(sec_match_col).lower().replace(" ", "")
                if sm in MATCH_OK:
                    self.w(f"{p}:{iri} skos:{MATCH_CANON[sm]} {ref_iri(sec_tgt)} .")

    # ---- meta-schema ----
    def header(self):
        p = self.p
        self.w(f"@prefix {p}:   <{self.ns}> .")
        self.w(f"@prefix ref:  <{REF_NS}> .")
        self.w("@prefix owl:  <http://www.w3.org/2002/07/owl#> .")
        self.w("@prefix rdf:  <http://www.w3.org/1999/02/22-rdf-syntax-ns#> .")
        self.w("@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .")
        self.w("@prefix xsd:  <http://www.w3.org/2001/XMLSchema#> .")
        self.w("@prefix skos: <http://www.w3.org/2004/02/skos/core#> .")
        self.w()
        self.w(f"<{BASE}{p}> a owl:Ontology ;")
        self.w(f"    owl:imports <{BASE}ref> ;")
        self.w(f"    rdfs:comment {lit(f'{p} vendor overlay. Deltas only: {p}: nodes plus SKOS mapping edges to the frozen ref: layer (imported, never restated). GENERATED from the capture workbook - do not hand-edit.')} .")
        self.w()
        self.w("# ============================================================")
        self.w(f"# {p}: META-SCHEMA")
        self.w("# ============================================================")
        self.w(f"{p}:DataObject a owl:Class ; rdfs:label {lit(p + ' Data Object')} .")
        self.w(f"{p}:ExternalReference a owl:Class ; rdfs:subClassOf {p}:DataObject ; "
               f"rdfs:label {lit(p + ' External Reference')} .")
        # fix D
        self.w(f"{p}:{self.OP_CLASS} a owl:Class ; rdfs:label {lit(p + ' Portal Operation (action verb)')} ; "
               f"rdfs:comment {lit('Verb axis. Named PortalOperation so it cannot collide with a vendor data object called Operation.')} .")
        self.w(f"{p}:operatesOn a owl:ObjectProperty ; rdfs:domain {p}:{self.OP_CLASS} ; "
               f"rdfs:range {p}:DataObject ; rdfs:label {lit('operatesOn')} .")
        self.w(f"{p}:verdict           a owl:AnnotationProperty .")
        self.w(f"{p}:sourceDoc         a owl:AnnotationProperty .")
        self.w(f"{p}:operationKind     a owl:AnnotationProperty .")
        self.w(f"{p}:confidence        a owl:DatatypeProperty ; rdfs:range xsd:decimal .")
        # fix E
        self.w(f"{p}:confirmsRefConcept a owl:AnnotationProperty ; "
               f"rdfs:comment {lit('Vendor independently confirms a ref: concept currently tagged Derived (Derived->Grounded signal).')} .")
        self.w(f"{p}:promotionNote      a owl:AnnotationProperty .")
        self.w(f"{p}:promotionEvidence  a owl:AnnotationProperty .")
        self.w(f"{p}:UnresolvedReference a owl:Class ; rdfs:subClassOf {p}:ExternalReference ; "
               f"rdfs:label {lit('Unresolved Reference')} ; "
               f"rdfs:comment {lit('Placeholder for a domain/range the master marks NEEDS VERIFICATION.')} .")
        # section 6: context-resolution sheets (optional, new-template only). Vendor-local
        # classes rather than ref:-namespaced ones - ref: is frozen/imported, never minted.
        self.w(f"{p}:ContextResolver a owl:Class ; rdfs:subClassOf {p}:DataObject ; "
               f"rdfs:label {lit(p + ' Context Resolver')} ; "
               f"rdfs:comment {lit('A table that resolves a target object at runtime from key field(s), rather than a static association.')} .")
        self.w(f"{p}:ContextField a owl:Class ; rdfs:label {lit(p + ' Context Field')} .")
        self.w(f"{p}:PrecedenceTier a owl:Class ; rdfs:label {lit(p + ' Context Precedence Tier')} .")
        self.w(f"{p}:resolvesTarget a owl:ObjectProperty ; rdfs:domain {p}:ContextResolver ; "
               f"rdfs:label {lit('resolvesTarget')} .")
        self.w(f"{p}:ofTable a owl:ObjectProperty ; rdfs:range {p}:ContextResolver ; "
               f"rdfs:label {lit('ofTable')} .")
        self.w(f"{p}:tableKind    a owl:AnnotationProperty .")
        self.w(f"{p}:isKey        a owl:DatatypeProperty ; rdfs:domain {p}:ContextField ; rdfs:range xsd:boolean .")
        self.w(f"{p}:isMandatory  a owl:DatatypeProperty ; rdfs:domain {p}:ContextField ; rdfs:range xsd:boolean .")
        self.w(f"{p}:fieldComment a owl:AnnotationProperty .")
        self.w(f"{p}:keyField a owl:ObjectProperty ; rdfs:domain {p}:PrecedenceTier ; "
               f"rdfs:range {p}:ContextField ; rdfs:label {lit('keyField')} .")
        self.w(f"{p}:tierOrder a owl:DatatypeProperty ; rdfs:domain {p}:PrecedenceTier ; "
               f"rdfs:range xsd:integer .")
        self.w()

    # ---- data objects ----
    def data_objects(self):
        sh, section = open_sheet(
            self.wb, ("1. Nodes", "1. Data Objects"),
            required=["Layer", "Name", "IRI local name", "Type"],
            optional=["Belongs to capability", "Affects data object", "ISA-95 area",
                      "Maps to neutral", "Match type", "VERDICT",
                      "Secondary maps to neutral", "Secondary match type",
                      "Confidence", "Source doc", "Notes"])
        p = self.p
        self.w("# ============================================================")
        self.w("# DATA OBJECTS")
        self.w("# ============================================================")
        self.w()
        n = 0
        for r in sh.rows():
            if norm(r.get("Layer")) != norm(p):
                continue
            name = r.str("Name")
            if not name:
                continue
            iri = cls_local(r.get("IRI local name") or name)
            self.define(iri, section)
            # remember BOTH the display name and the IRI so later sheets can refer to
            # an object by either ("Parametric Data Definition (CDO)" -> ParametricDataDefinition)
            self.name_to_iri[norm(name)] = iri
            self.name_to_iri[norm(iri)] = iri
            typ = r.str("Type")
            parent = f"{p}:ExternalReference" if "external" in typ.lower() else f"{p}:DataObject"

            parts = [f"{p}:{iri} a owl:Class ; rdfs:subClassOf {parent} ; rdfs:label {lit(name)}"]
            v = self.verdict_for(sh, r)
            if v:
                parts.append(f"    {p}:verdict {lit(v)}")
            if sh.has("Source doc") and r.str("Source doc"):
                parts.append(f"    {p}:sourceDoc {lit(r.str('Source doc'))}")
            if sh.has("Confidence"):                      # fix A/B: optional, numeric only
                d = as_decimal(r.get("Confidence"))
                if d is not None:
                    parts.append(f"    {p}:confidence {d}")
            if sh.has("Notes") and r.str("Notes"):
                parts.append(f"    rdfs:comment {lit(r.str('Notes'))}")
            self.w(" ;\n".join(parts) + " .")

            self.emit_mapping_edges(sh, r, iri)
            self.w()
            n += 1
        self.counts["data objects"] = n

    def resolve_class(self, name):
        """
        Map a Domain/Range/operatesOn cell to a DEFINED class IRI.
        Sheets refer to objects by DISPLAY name ("Parametric Data Definition (CDO)")
        while the class is minted from the 'IRI local name' column
        ("ParametricDataDefinition"). Transforming the display name directly produced
        dangling IRIs, so resolve through the Data Objects map first.
        """
        raw = str(name).strip()
        # Masters annotate ranges parenthetically: "Area (self)", "Flow (self of parent)",
        # "NEEDS VERIFICATION (likely Step)". The parenthetical is a note, not part of the
        # name - strip it before resolving, or we mint dangling IRIs like Areaself.
        base = re.sub(r"\s*\([^)]*\)\s*$", "", raw).strip()
        for cand in (norm(raw), norm(base)):
            if cand in self.name_to_iri:
                return self.name_to_iri[cand]
        # A cell may list SEVERAL alternative targets ("Area | Material | Resource").
        # RDFS rdfs:range is conjunctive, so emitting a fused IRI is wrong and emitting
        # all of them would over-constrain. Resolve to the first known class and record
        # the alternatives as a warning.
        if re.search(r"[|/]| or ", base, re.I):
            alts = [a.strip() for a in re.split(r"[|/]| or ", base, flags=re.I) if a.strip()]
            for a in alts:
                if norm(a) in self.name_to_iri:
                    self.warnings.append(
                        f"multi-target range {raw!r} resolved to {a!r} "
                        f"(alternatives not emitted; rdfs:range is conjunctive)")
                    return self.name_to_iri[norm(a)]
        if base.lower() in ("string", "text", "integer", "boolean", "decimal", "datetime"):
            self.warnings.append(
                f"object-property range looks like a datatype: {raw!r} -> placeholder")
            return "UnresolvedReference"
        if base.upper().startswith("NEEDS VERIFICATION"):
            self.warnings.append(
                f"unresolved range/domain kept as placeholder: {raw!r}")
            return "UnresolvedReference"
        return cls_local(base or raw)

    # ---- enums & lookups (fix F: keep allowed values) ----
    def enums(self):
        sh = Sheet(self.wb, "4. Enums & Lookups",
                   required=["Set name", "Kind", "Allowed values", "Layer"],
                   optional=["Used by", "Notes"])
        p = self.p
        self.w("# ============================================================")
        self.w("# ENUMS & LOOKUPS")
        self.w("# ============================================================")
        self.w()
        n = 0
        for r in sh.rows():
            setname = r.str("Set name")
            if not setname:
                continue
            iri = cls_local(setname)
            self.define(iri, "4. Enums & Lookups")
            kind = r.str("Kind").lower()
            values = r.str("Allowed values")
            notes = r.str("Notes")
            vals = [v.strip() for v in re.split(r"[;|]", values) if v.strip()] if values else []
            extensible = "extensible" in values.lower() or kind == "lookup"

            if kind == "enum" and vals and not extensible:
                vlist = " ".join(lit(v) for v in vals)
                parts = [f"{p}:{iri} a rdfs:Datatype ; rdfs:label {lit(setname)}",
                         f"    owl:equivalentClass [ a rdfs:Datatype ; owl:oneOf ( {vlist} ) ]"]
                if notes:
                    parts.append(f"    rdfs:comment {lit(notes)}")
                self.w(" ;\n".join(parts) + " .")
            else:
                # extensible lookup: ConceptScheme + known members, without implying closure
                parts = [f"{p}:{iri} a skos:ConceptScheme ; rdfs:label {lit(setname)}"]
                if notes:
                    parts.append(f"    rdfs:comment {lit(notes)}")
                self.w(" ;\n".join(parts) + " .")
                for v in vals:
                    if "extensible" in v.lower():
                        continue
                    member = f"{iri}_{local(v)}"
                    self.w(f"{p}:{member} a skos:Concept ; skos:inScheme {p}:{iri} ; "
                           f"skos:prefLabel {lit(v)} .")
            self.w()
            n += 1
        self.counts["enum/lookup sets"] = n

    # ---- attributes (fix C: qualify only on collision) ----
    def attributes(self):
        sh, section = open_sheet(
            self.wb, ("2. Data Properties", "2. Attributes"),
            required=["Layer", "Data object", "Attribute", "IRI local name", "Datatype"],
            legacy_required=["Data object", "Attribute", "IRI local name", "Datatype"],
            optional=["Value kind", "Enum/Lookup set", "Cardinality", "Conditional on",
                      "Maps to neutral attribute", "Match type", "VERDICT",
                      "Secondary maps to neutral", "Secondary match type",
                      "Confidence", "Source doc", "Notes"])
        p = self.p
        has_layer = sh.has("Layer")

        def in_scope(r):
            return not has_layer or norm(r.get("Layer")) == norm(p)

        # pre-scan to find local names used by more than one owning class
        seen = {}
        for r in sh.rows():
            if not in_scope(r):
                continue
            dobj, iri = r.str("Data object"), local(r.get("IRI local name") or r.str("Attribute"))
            if dobj and iri:
                seen.setdefault(iri, set()).add(dobj)
        collide = {k for k, v in seen.items() if len(v) > 1}

        self.w("# ============================================================")
        self.w("# ATTRIBUTES (data properties)")
        self.w("# ============================================================")
        self.w()
        n = 0
        for r in sh.rows():
            if not in_scope(r):
                continue
            dobj = r.str("Data object")
            if not dobj:
                continue
            attr = r.str("Attribute")
            base = local(r.get("IRI local name") or attr)
            # only auto-qualify when the hand-authored name is ambiguous
            iri = (local(dobj) + base[:1].upper() + base[1:]) if base in collide else base
            self.define(iri, section)

            vkind = r.str("Value kind").lower()
            eset = r.str("Enum/Lookup set")
            # If the "lookup set" is actually a defined DATA OBJECT (e.g. Resource.Factory
            # -> the Factory class), this is an object reference, not a literal. Emitting it
            # as a DatatypeProperty with a class range is invalid; emit an ObjectProperty.
            as_object_ref = bool(eset) and norm(eset) in self.name_to_iri
            if as_object_ref:
                rng = f"{p}:{self.resolve_class(eset)}"
            elif vkind in ("enum", "lookup") and eset:
                rng = f"{p}:{cls_local(eset)}"
            else:
                rng = DTYPE.get(r.str("Datatype").lower(), "xsd:string")

            ptype = "owl:ObjectProperty" if as_object_ref else "owl:DatatypeProperty"
            parts = [f"{p}:{iri} a {ptype} ; rdfs:label {lit(attr)} ; "
                     f"rdfs:domain {p}:{self.resolve_class(dobj)}",
                     f"    rdfs:range {rng}"]
            v = self.verdict_for(sh, r, mapping_col="Maps to neutral attribute")
            if v:
                parts.append(f"    {p}:verdict {lit(v)}")
            if sh.has("Confidence"):
                d = as_decimal(r.get("Confidence"))
                if d is not None:
                    parts.append(f"    {p}:confidence {d}")
            bits = []
            if r.str("Cardinality"):
                bits.append(f"cardinality {r.str('Cardinality')}")
            if r.str("Conditional on"):
                bits.append(f"conditional on {r.str('Conditional on')}")
            if r.str("Notes"):
                bits.append(r.str("Notes"))
            if bits:
                parts.append(f"    rdfs:comment {lit('; '.join(bits))}")
            self.w(" ;\n".join(parts) + " .")

            self.emit_mapping_edges(sh, r, iri, mapping_col="Maps to neutral attribute")
            self.w()
            n += 1
        self.counts["attributes"] = n

    # ---- object properties (fix F: carry Notes) ----
    def object_properties(self):
        # sheet name is unchanged between templates; only require "Layer" when the
        # workbook's copy of this sheet actually has that column (new template does,
        # the legacy CM master's "3. Object Properties" does not).
        has_layer_col = "layer" in sheet_header_names(self.wb, "3. Object Properties")
        sh = Sheet(self.wb, "3. Object Properties",
                   required=["Domain", "Property", "IRI local name", "Range"]
                            + (["Layer"] if has_layer_col else []),
                   optional=["Layer", "Cardinality", "Conditional on", "Maps to neutral",
                             "Match type", "Secondary maps to neutral", "Secondary match type",
                             "Confidence", "VERDICT", "Source doc", "Notes"])
        p = self.p
        has_layer = sh.has("Layer")

        def in_scope(r):
            return not has_layer or norm(r.get("Layer")) == norm(p)

        seen = {}
        for r in sh.rows():
            if not in_scope(r):
                continue
            dom, iri = r.str("Domain"), local(r.get("IRI local name") or r.str("Property"))
            if dom and iri:
                seen.setdefault(iri, set()).add(dom)
        collide = {k for k, v in seen.items() if len(v) > 1}

        self.w("# ============================================================")
        self.w("# OBJECT PROPERTIES")
        self.w("# ============================================================")
        self.w()
        n = 0
        for r in sh.rows():
            if not in_scope(r):
                continue
            dom = r.str("Domain")
            if not dom:
                continue
            prop = r.str("Property")
            base = local(r.get("IRI local name") or prop)
            iri = (local(dom) + base[:1].upper() + base[1:]) if base in collide else base
            self.define(iri, "3. Object Properties")

            parts = [f"{p}:{iri} a owl:ObjectProperty ; rdfs:label {lit(prop)} ; "
                     f"rdfs:domain {p}:{self.resolve_class(dom)}",
                     f"    rdfs:range {p}:{self.resolve_class(r.str('Range'))}"]
            bits = []
            if r.str("Cardinality"):
                bits.append(f"cardinality {r.str('Cardinality')}")
            if r.str("Conditional on"):
                bits.append(f"conditional on {r.str('Conditional on')}")
            if sh.has("Notes") and r.str("Notes"):
                bits.append(r.str("Notes"))
            if bits:
                parts.append(f"    rdfs:comment {lit('; '.join(bits))}")
            v = self.verdict_for(sh, r)
            if v:
                parts.append(f"    {p}:verdict {lit(v)}")
            if sh.has("Confidence"):
                d = as_decimal(r.get("Confidence"))
                if d is not None:
                    parts.append(f"    {p}:confidence {d}")
            self.w(" ;\n".join(parts) + " .")

            self.emit_mapping_edges(sh, r, iri)
            self.w()
            n += 1
        self.counts["object properties"] = n

    # ---- operations (fix D: PortalOperation) ----
    def operations(self):
        if "5. Operations" not in self.wb.sheetnames:
            self.counts["operations"] = 0
            return
        sh = Sheet(self.wb, "5. Operations", required=["Label"],
                   optional=["Operation IRI", "Kind", "operatesOn", "Maps to (peer vendor)",
                             "Match type", "Confidence", "VERDICT", "Source doc", "Notes"])
        p = self.p
        self.w("# ============================================================")
        self.w("# OPERATIONS (verb axis; lateral vendor->peer mapping)")
        self.w("# ============================================================")
        self.w()
        # operatesOn column name varies by master (e.g. "operatesOn (cm:)")
        op_col = next((h for h in sh.headers if h.startswith("operateson")), None)
        n = 0
        for r in sh.rows():
            label = r.str("Label")
            if not label:
                continue
            src = r.str("Operation IRI") or label
            iri = cls_local(src)
            self.define(iri, "5. Operations")
            parts = [f"{p}:{iri} a {p}:{self.OP_CLASS} ; rdfs:label {lit(label)}"]
            if op_col:
                tgt = str(r._row[sh.headers[op_col]] or "").strip()
                if tgt:
                    parts.append(f"    {p}:operatesOn {p}:{self.resolve_class(tgt)}")
            for col, prop in (("Kind", "operationKind"), ("VERDICT", "verdict"),
                              ("Source doc", "sourceDoc")):
                if sh.has(col) and r.str(col):
                    parts.append(f"    {p}:{prop} {lit(r.str(col))}")
            if sh.has("Confidence"):
                d = as_decimal(r.get("Confidence"))
                if d is not None:
                    parts.append(f"    {p}:confidence {d}")
            if sh.has("Notes") and r.str("Notes"):
                parts.append(f"    rdfs:comment {lit(r.str('Notes'))}")
            self.w(" ;\n".join(parts) + " .")
            self.w()
            n += 1
        self.counts["operations"] = n

    # ---- promotions (fix E) ----
    def promotions(self):
        name = next((s for s in self.wb.sheetnames if "promotion" in s.lower()), None)
        if not name:
            self.counts["promotions"] = 0
            return
        sh = Sheet(self.wb, name, required=[],
                   optional=["Subject", "Vendor concept", "Ref concept", "Target",
                             "Evidence", "Note", "Notes"])
        p = self.p
        self.w("# ============================================================")
        self.w("# PROMOTIONS (vendor confirms a Derived ref: concept)")
        self.w("# ============================================================")
        self.w()
        # column synonyms - masters spell these differently
        SUBJ = ("subject", "vendor concept", "opcenter evidence", "vendor evidence",
                "evidence", "source screen")
        TGT  = ("ref concept", "target", "ref: target", "ref target")
        NOTE = ("notes", "note", "recommendation", "rationale")
        subj_col = next((h for h in SUBJ if h in sh.headers), None)
        tgt_col  = next((h for h in TGT  if h in sh.headers), None)
        note_col = next((h for h in NOTE if h in sh.headers), None)
        if not tgt_col:
            raise BuildError(
                f"Sheet '{name}' has no recognisable ref-target column. "
                f"Looked for {list(TGT)}; found {sorted(sh.headers)}.")
        n = 0
        if subj_col and tgt_col:
            for r in sh.rows():
                t = str(r._row[sh.headers[tgt_col]] or "").strip()
                if not t:
                    continue
                raw_s = str(r._row[sh.headers[subj_col]] or "").strip() if subj_col else ""
                # evidence text is often prose ("opc:Container.Level (Batch...)").
                # take a vendor IRI if one is present, else anchor on the ref target.
                m = re.search(rf"{p}:([A-Za-z0-9_.]+)", raw_s)
                subj = cls_local(m.group(1).split(".")[0]) if m else cls_local(local(t))
                parts = [f"{p}:{subj} {p}:confirmsRefConcept {ref_iri(t)}"]
                if raw_s:
                    parts.append(f"    {p}:promotionEvidence {lit(raw_s)}")
                if note_col:
                    note = str(r._row[sh.headers[note_col]] or "").strip()
                    if note:
                        parts.append(f"    {p}:promotionNote {lit(note)}")
                self.w(" ;\n".join(parts) + " .")
                n += 1
            self.w()
        self.counts["promotions"] = n

    # ---- context sheets (section 6, new-template only) ----
    def context_tables(self):
        """
        Sheet '5. Context Tables' is absent from the legacy CM master -> skip silently.
        Groups rows by Table name; the FIRST row of each group supplies the table-level
        columns (Resolves (target), Table kind, Maps to neutral, Match type, VERDICT,
        Source doc, Notes) - later rows for the same table only contribute a field.
        """
        if "5. Context Tables" not in self.wb.sheetnames:
            self.counts["context tables"] = 0
            return
        sh = Sheet(self.wb, "5. Context Tables",
                   required=["Layer", "Table name", "Field"],
                   optional=["Table kind", "Resolves (target)", "Is key", "Is mandatory",
                             "Field comment", "Maps to neutral", "Match type", "VERDICT",
                             "Source doc", "Notes"])
        p = self.p
        self.w("# ============================================================")
        self.w("# CONTEXT TABLES (runtime resolution, not static association)")
        self.w("# ============================================================")
        self.w()
        tables, order = {}, []
        for r in sh.rows():
            if norm(r.get("Layer")) != norm(p):
                continue
            tname = r.str("Table name")
            if not tname:
                continue
            if tname not in tables:
                tables[tname] = {"iri": cls_local(tname), "first": r, "fields": []}
                order.append(tname)
            tables[tname]["fields"].append(r)

        n = 0
        for tname in order:
            t = tables[tname]
            iri, first = t["iri"], t["first"]
            self.define(iri, "5. Context Tables")
            parts = [f"{p}:{iri} a {p}:ContextResolver ; rdfs:label {lit(tname)}"]
            target = first.str("Resolves (target)")
            if target:
                parts.append(f"    {p}:resolvesTarget {p}:{self.resolve_class(target)}")
            if sh.has("Table kind") and first.str("Table kind"):
                parts.append(f"    {p}:tableKind {lit(first.str('Table kind'))}")
            v = self.verdict_for(sh, first)
            if v:
                parts.append(f"    {p}:verdict {lit(v)}")
            if sh.has("Source doc") and first.str("Source doc"):
                parts.append(f"    {p}:sourceDoc {lit(first.str('Source doc'))}")
            if sh.has("Notes") and first.str("Notes"):
                parts.append(f"    rdfs:comment {lit(first.str('Notes'))}")
            self.w(" ;\n".join(parts) + " .")
            self.emit_mapping_edges(sh, first, iri)
            self.w()

            for fi, fr in enumerate(t["fields"], start=1):
                fname = fr.str("Field")
                if not fname:
                    continue
                field_iri = f"{iri}_field{fi}"
                self.field_iris[(tname, norm(fname))] = field_iri
                fparts = [f"{p}:{field_iri} a {p}:ContextField ; rdfs:label {lit(fname)} ; "
                          f"{p}:ofTable {p}:{iri}",
                          f"    {p}:isKey {yes_no(fr.get('Is key'))}",
                          f"    {p}:isMandatory {yes_no(fr.get('Is mandatory'))}"]
                if sh.has("Field comment") and fr.str("Field comment"):
                    fparts.append(f"    {p}:fieldComment {lit(fr.str('Field comment'))}")
                self.w(" ;\n".join(fparts) + " .")
            self.w()
            n += 1
        self.counts["context tables"] = n

    def context_precedence(self):
        """
        Sheet '6. Context Precedence' is absent from the legacy CM master -> skip
        silently. One PrecedenceTier node per row, ordered by the Order column (numeric
        when it parses, else encounter order), linked to its table and to the
        ContextField node of each ';'-separated field in Key combination (matched by
        field DISPLAY NAME against fields emitted in context_tables() for that same
        table). An unresolvable field name is a warning, not a fatal error - the tier is
        still emitted with a comment recording the raw name.
        """
        if "6. Context Precedence" not in self.wb.sheetnames:
            self.counts["context precedence tiers"] = 0
            return
        sh = Sheet(self.wb, "6. Context Precedence",
                   required=["Layer", "Table name", "Order", "Key combination"],
                   optional=["Notes"])
        p = self.p
        self.w("# ============================================================")
        self.w("# CONTEXT PRECEDENCE")
        self.w("# ============================================================")
        self.w()
        groups, order_list = {}, []
        for r in sh.rows():
            if norm(r.get("Layer")) != norm(p):
                continue
            tname = r.str("Table name")
            if not tname:
                continue
            if tname not in groups:
                groups[tname] = []
                order_list.append(tname)
            groups[tname].append(r)

        def sort_key(r):
            d = as_decimal(r.get("Order"))
            return float(d) if d is not None else float("inf")

        n = 0
        for tname in order_list:
            tiri = cls_local(tname)
            for r in sorted(groups[tname], key=sort_key):
                order_val = r.str("Order")
                n += 1
                tier_iri = f"{tiri}_precedence{local(order_val) or n}"
                parts = [f"{p}:{tier_iri} a {p}:PrecedenceTier ; {p}:ofTable {p}:{tiri}"]
                d = as_decimal(order_val)
                if d is not None:
                    parts.append(f"    {p}:tierOrder {d}")
                if r.str("Notes"):
                    parts.append(f"    rdfs:comment {lit(r.str('Notes'))}")
                self.w(" ;\n".join(parts) + " .")
                combo = [c.strip() for c in r.str("Key combination").split(";") if c.strip()]
                for field_name in combo:
                    key = (tname, norm(field_name))
                    if key in self.field_iris:
                        self.w(f"{p}:{tier_iri} {p}:keyField {p}:{self.field_iris[key]} .")
                    else:
                        self.warnings.append(
                            f"WARNING: Context Precedence table {tname!r} order "
                            f"{order_val!r} references field {field_name!r} not found "
                            f"among that table's Context Tables fields.")
                        self.w(f"# unresolved key field for {p}:{tier_iri}: {field_name}")
                self.w()
        self.counts["context precedence tiers"] = n

    def generate(self):
        self.check_template_version()
        self.validate_rows()
        self.header()
        self.data_objects()
        self.enums()
        self.attributes()
        self.object_properties()
        self.operations()
        self.promotions()
        self.context_tables()
        self.context_precedence()
        # fix D: meta IRIs must not collide with minted IRIs
        clash = self.meta_iris & set(self.defined)
        if clash:
            raise BuildError(
                f"Meta-schema IRI(s) {sorted(clash)} collide with names minted from the "
                f"master. Rename in the master or adjust the meta-schema.")
        return "\n".join(self.out) + "\n"


# ---------------------------------------------------------------------------
# Validation pass (fix G)
# ---------------------------------------------------------------------------
def validate(ttl, prefix, counts, warnings=()):
    problems = []

    # 1. re-parse with pyoxigraph
    try:
        import pyoxigraph as ox
        store = ox.Store()
        store.load(ttl.encode("utf-8"), format=ox.RdfFormat.TURTLE,
                   to_graph=ox.NamedNode("urn:validate"))
        triples = sum(1 for _ in store.quads_for_pattern(
            None, None, None, ox.NamedNode("urn:validate")))
    except ImportError:
        triples = None
        problems.append("WARN: pyoxigraph not installed - skipped parse check.")
    except Exception as e:
        problems.append(f"FAIL: output is not valid Turtle: {e}")
        triples = None

    ns = f"{BASE}{prefix}#"

    # 2. mapping targets must be in ref:
    for m in re.finditer(r"skos:\w+Match\s+(\S+)\s*\.", ttl):
        tgt = m.group(1)
        if not tgt.startswith("ref:"):
            problems.append(f"FAIL: SKOS mapping target outside ref: -> {tgt}")

    # 3. ref: must never be a subject
    for line in ttl.splitlines():
        if re.match(r"^\s*ref:\S+\s+(a|rdfs:|owl:)", line):
            problems.append(f"FAIL: ref: IRI used as a subject -> {line.strip()[:70]}")

    # 4. datatype property ranges must be datatypes, not classes
    declared_datatypes = set(re.findall(rf"{prefix}:(\w+) a rdfs:Datatype", ttl))
    for m in re.finditer(rf"{prefix}:(\w+) a owl:DatatypeProperty[^.]*?rdfs:range\s+(\S+)", ttl, re.S):
        prop, rng = m.groups()
        if rng.startswith(f"{prefix}:"):
            local_rng = rng.split(":", 1)[1]
            if local_rng not in declared_datatypes and not rng.startswith("xsd:"):
                # ConceptScheme ranges are an accepted modelling choice for lookups
                if f"{prefix}:{local_rng} a skos:ConceptScheme" not in ttl:
                    problems.append(
                        f"FAIL: datatype property {prefix}:{prop} has class range {rng}")

    # 5. no bare (unquoted) annotation objects.
    #    Skip the meta-schema declarations themselves (":verdict a owl:AnnotationProperty"),
    #    which legitimately have a non-literal object.
    for m in re.finditer(rf"{prefix}:(verdict|sourceDoc|operationKind)\s+([^\s\"<][^\s;.]*)", ttl):
        obj = m.group(2)
        if obj == "a" or obj.startswith("owl:") or obj.startswith("rdfs:"):
            continue
        problems.append(f"FAIL: unquoted literal -> {prefix}:{m.group(1)} {obj}")

    # 6. NON-EMPTINESS (a silently empty section is the worst failure mode:
    #    the build "succeeds" and produces an overlay with no link to ref: at all).
    if counts.get("data objects", 0) == 0:
        problems.append("FAIL: zero data objects emitted - check the Layer column "
                        "values match the --prefix, and that sheet '1. Data Objects' "
                        "has rows for this vendor.")
    n_map = len(re.findall(r"skos:\w+Match", ttl))
    if n_map == 0:
        problems.append("FAIL: zero SKOS mapping edges emitted - the overlay would have "
                        "no link to ref: and is unusable for comparison.")

    # 7. DANGLING IRIs: anything referenced by domain/range/operatesOn must be defined.
    defined_iris = set(re.findall(rf"^{prefix}:(\w+) a ", ttl, re.M))
    referenced = set()
    for pat in (rf"rdfs:domain\s+{prefix}:(\w+)", rf"rdfs:range\s+{prefix}:(\w+)",
                rf"{prefix}:operatesOn\s+{prefix}:(\w+)"):
        referenced |= set(re.findall(pat, ttl))
    dangling = sorted(referenced - defined_iris)
    if dangling:
        problems.append(
            f"FAIL: {len(dangling)} IRI(s) referenced but never defined: "
            f"{dangling[:8]}{' ...' if len(dangling) > 8 else ''}")

    print("\n--- generation summary ---")
    for k, v in counts.items():
        print(f"  {k:22} {v}")
    print(f"  {'skos mapping edges':22} {n_map}")
    if triples is not None:
        print(f"  {'triples parsed':22} {triples}")
    if warnings:
        print("\n--- warnings ---")
        for w in warnings:
            print("  " + w)
    if problems:
        print("\n--- validation problems ---")
        for p in problems:
            print("  " + p)
    else:
        print("\n  validation: all checks passed")
    return [p for p in problems if p.startswith("FAIL")]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--prefix", required=True, help="vendor prefix, e.g. opc")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    try:
        gen = Generator(args.workbook, args.prefix)
        ttl = gen.generate()
        gen.wb.close()
    except BuildError as e:
        print(f"BUILD FAILED: {e}", file=sys.stderr)
        sys.exit(1)

    failures = validate(ttl, args.prefix, gen.counts, gen.warnings)
    if failures:
        print(f"\nBUILD FAILED: {len(failures)} validation failure(s); file not written.",
              file=sys.stderr)
        sys.exit(1)

    with open(args.out, "w", encoding="utf-8") as f:
        f.write(ttl)
    print(f"\nwrote {args.out}")


if __name__ == "__main__":
    main()
